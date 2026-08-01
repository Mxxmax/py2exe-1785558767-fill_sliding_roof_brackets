#!/usr/bin/env python3
"""Recursively find supplier workbooks and fill a summary .xlsm by label."""

from __future__ import annotations

import argparse
import json
import posixpath
import re
import sys
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import et_xmlfile  # noqa: F401 — openpyxl 的隐式依赖，确保 PyInstaller 打包

from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel
from lxml import etree as ET


SUPPLIER_COLUMNS = range(5, 14)  # E:M
DEFAULT_ROWS = {"date": 15, "material_basis": 16, "material": 17,
                "purchased": 18, "production": 19, "overhead_profit": 20,
                "packaging": 21, "freight": 24, "tooling": 30}

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"


def _sheet_xml_paths(book: zipfile.ZipFile) -> dict[str, str]:
    workbook = ET.fromstring(book.read("xl/workbook.xml"))
    rels = ET.fromstring(book.read("xl/_rels/workbook.xml.rels"))
    targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall(f"{{{PKG_REL_NS}}}Relationship")
    }
    result = {}
    sheets = workbook.find(f"{{{MAIN_NS}}}sheets")
    for sheet in sheets if sheets is not None else []:
        rel_id = sheet.attrib[f"{{{DOC_REL_NS}}}id"]
        target = targets[rel_id].lstrip("/")
        if not target.startswith("xl/"):
            target = posixpath.normpath(posixpath.join("xl", target))
        result[sheet.attrib["name"]] = target
    return result


def _set_xml_cell(root: ET.Element, coordinate: str, value: Any) -> None:
    cell = root.find(f".//{{{MAIN_NS}}}c[@r='{coordinate}']")
    if cell is None:
        raise LookupError(f"目标工作簿中不存在单元格 {coordinate}")
    for child in list(cell):
        if child.tag in {f"{{{MAIN_NS}}}f", f"{{{MAIN_NS}}}v", f"{{{MAIN_NS}}}is"}:
            cell.remove(child)
    if isinstance(value, datetime):
        cell.set("t", "n")
        ET.SubElement(cell, f"{{{MAIN_NS}}}v").text = str(to_excel(value))
    elif isinstance(value, bool):
        cell.set("t", "b")
        ET.SubElement(cell, f"{{{MAIN_NS}}}v").text = "1" if value else "0"
    elif isinstance(value, (int, float)):
        cell.set("t", "n")
        ET.SubElement(cell, f"{{{MAIN_NS}}}v").text = format(value, ".15g")
    elif value is None:
        cell.attrib.pop("t", None)
    else:
        cell.set("t", "inlineStr")
        inline = ET.SubElement(cell, f"{{{MAIN_NS}}}is")
        text = ET.SubElement(inline, f"{{{MAIN_NS}}}t")
        text.text = str(value)


def write_preserving_excel_features(source: Path, output: Path,
                                    updates: dict[str, dict[str, Any]]) -> None:
    """Patch target cells and force Excel to rebuild formulas on open."""
    with zipfile.ZipFile(source, "r") as src:
        sheet_paths = _sheet_xml_paths(src)
        patched: dict[str, bytes] = {}
        for sheet_name, cells in updates.items():
            path = sheet_paths[sheet_name]
            root = ET.fromstring(src.read(path))
            for coordinate, value in cells.items():
                _set_xml_cell(root, coordinate, value)
            patched[path] = ET.tostring(root, encoding="utf-8", xml_declaration=True)

        # Updated inputs invalidate the workbook's old calculation chain.
        workbook_root = ET.fromstring(src.read("xl/workbook.xml"))
        calc_pr = workbook_root.find(f"{{{MAIN_NS}}}calcPr")
        if calc_pr is None:
            calc_pr = ET.SubElement(workbook_root, f"{{{MAIN_NS}}}calcPr")
        calc_pr.attrib.update({
            "calcId": "0",
            "calcMode": "auto",
            "calcCompleted": "0",
            "fullCalcOnLoad": "1",
            "forceFullCalc": "1",
            "calcOnSave": "1",
        })
        patched["xl/workbook.xml"] = ET.tostring(
            workbook_root, encoding="utf-8", xml_declaration=True)

        rels_path = "xl/_rels/workbook.xml.rels"
        rels_root = ET.fromstring(src.read(rels_path))
        for relationship in list(rels_root):
            if relationship.attrib.get("Type", "").endswith("/calcChain"):
                rels_root.remove(relationship)
        patched[rels_path] = ET.tostring(
            rels_root, encoding="utf-8", xml_declaration=True)

        content_types_path = "[Content_Types].xml"
        content_types_root = ET.fromstring(src.read(content_types_path))
        for override in list(content_types_root):
            if override.attrib.get("PartName") == "/xl/calcChain.xml":
                content_types_root.remove(override)
        patched[content_types_path] = ET.tostring(
            content_types_root, encoding="utf-8", xml_declaration=True)

        with zipfile.ZipFile(output, "w") as dst:
            for info in src.infolist():
                if info.filename == "xl/calcChain.xml":
                    continue
                dst.writestr(info, patched.get(info.filename, src.read(info.filename)))


def norm(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9\u0370-\u03ff\u4e00-\u9fff]+", "", str(value).lower())


def number(value: Any) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return 0.0
    return float(value)


def contains(cell_value: Any, *parts: str) -> bool:
    text = norm(cell_value)
    return any(norm(p) in text for p in parts)


def find_label(ws, *parts: str, min_column: int = 1):
    for row in ws.iter_rows():
        for cell in row:
            if cell.column >= min_column and contains(cell.value, *parts):
                return cell
    raise LookupError(f"找不到标签: {parts}")


def find_exact_label(ws, *labels: str, min_column: int = 1):
    wanted = {norm(label) for label in labels}
    for row in ws.iter_rows():
        for cell in row:
            if cell.column >= min_column and norm(cell.value) in wanted:
                return cell
    raise LookupError(f"找不到精确标签: {labels}")


def find_rightmost_label(ws, *parts: str, exact: bool = False):
    """同一标签出现两次时，最右侧的一组是欧元报价区域。"""
    wanted = {norm(part) for part in parts}
    matches = []
    for row in ws.iter_rows():
        for cell in row:
            text = norm(cell.value)
            matched = text in wanted if exact else any(part in text for part in wanted)
            if matched:
                matches.append(cell)
    if not matches:
        raise LookupError(f"找不到标签: {parts}")
    return max(matches, key=lambda cell: (cell.column, -cell.row))


def value_by_exact_label(ws, data_ws, labels: Iterable[str]):
    wanted = {norm(label) for label in labels}
    label = find_rightmost_label(ws, *wanted, exact=True)
    return next_numeric_right(data_ws, data_ws.cell(label.row, label.column), ws.max_column)


def value_at_label_and_column(ws, data_ws, labels: Iterable[str], result_column: int):
    label = find_rightmost_label(ws, *labels, exact=True)
    value = data_ws.cell(label.row, result_column).value
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        raise LookupError(f"标签 {label.coordinate} 在结果列 {result_column} 没有数值")
    return value


def next_numeric_right(ws, cell, max_columns: int = 20):
    for col in range(cell.column + 1, min(ws.max_column, cell.column + max_columns) + 1):
        value = ws.cell(cell.row, col).value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return value
        if isinstance(value, datetime):
            return value
    raise LookupError(f"标签 {cell.coordinate} 右侧没有数值")


def next_value_right(ws, cell, max_columns: int = 20):
    for col in range(cell.column + 1, min(ws.max_column, cell.column + max_columns) + 1):
        value = ws.cell(cell.row, col).value
        if value not in (None, ""):
            return value
    raise LookupError(f"标签 {cell.coordinate} 右侧没有值")


def row_value(ws, data_ws, *parts: str, preferred_cols=(41, 40, 23, 22, 27)):
    label = find_label(ws, *parts)
    for col in preferred_cols:
        value = data_ws.cell(label.row, col).value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return value
    return next_numeric_right(data_ws, data_ws.cell(label.row, label.column))


def section_rows(ws, header_parts: Iterable[str], end_parts: Iterable[str], min_column=1):
    start = find_label(ws, *header_parts, min_column=min_column).row
    end = find_label(ws, *end_parts, min_column=min_column).row
    if end <= start:
        raise LookupError(f"区间顺序异常: {header_parts} -> {end_parts}")
    return range(start + 1, end)


def max_material_price(ws, data_ws):
    header = find_rightmost_label(ws, "raw material price")
    total = find_rightmost_label(ws, "Σ 1", exact=True)
    if total.row <= header.row:
        raise LookupError("欧元区 raw material price 与 Σ 1 的行顺序异常")
    rows = range(header.row + 1, total.row)
    values = [number(data_ws.cell(r, header.column).value) for r in rows]
    values = [v for v in values if v != 0]
    return max(values) if values else 0


def select_quote_sheet(formulas, values, country: str):
    """Select the quote sheet by labels instead of assuming it is sheet 1."""
    required = ("quotation date", "raw material price", "production costs", "packaging costs")
    candidates = []
    for index, ws in enumerate(formulas.worksheets):
        texts = [norm(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
        score = sum(any(norm(label) in text for text in texts) for label in required)
        candidates.append((score, -index, ws))
    score, _, ws = max(candidates, key=lambda item: (item[0], item[1]))
    if score < 3:
        raise LookupError(f"{country} 报价单中找不到包含关键标签的 sheet")
    return ws, values[ws.title]


def extract_china_values(ws, ds) -> dict[str, Any]:
    date_label = find_rightmost_label(ws, "quotation date")
    date_value = next_value_right(ds, ds.cell(date_label.row, date_label.column), 20)
    material_label = find_rightmost_label(ws, "Σ 1", exact=True)
    purchased_label = find_rightmost_label(ws, "Σ 2", exact=True)
    material = next_numeric_right(ds, ds.cell(material_label.row, material_label.column), ws.max_column)
    purchased = next_numeric_right(ds, ds.cell(purchased_label.row, purchased_label.column), ws.max_column)
    material_result_col = find_rightmost_label(ws, "cost of material", exact=True).column
    production_result_col = find_rightmost_label(ws, "production costs", exact=True).column
    production = value_at_label_and_column(
        ws, ds, ("manufaction costs", "manufacturing costs"), production_result_col)
    setup = value_at_label_and_column(ws, ds, ("set up costs",), production_result_col)
    material_profit = value_at_label_and_column(
        ws, ds, ("+ additional profit on material",), material_result_col)
    overhead = value_at_label_and_column(ws, ds, ("+ overhead",), production_result_col)
    process_profit = value_at_label_and_column(
        ws, ds, ("+ profit on manufcturing process", "+ profit on manufacturing process"),
        production_result_col)
    packaging_total = find_rightmost_label(ws, "costs2, incl. packaging")
    freight_total = find_rightmost_label(ws, "FOB costs3")
    packaging_header = find_rightmost_label(ws, "packaging costs", exact=True)
    freight_header = find_rightmost_label(ws, "shipping costs", exact=True)
    packaging = number(ds.cell(packaging_total.row, packaging_header.column).value)
    freight = number(ds.cell(freight_total.row, freight_header.column).value)
    tooling = value_at_label_and_column(ws, ds, ("Tooling cost",), production_result_col)

    return {
        "date": date_value,
        "material_basis": max_material_price(ws, ds),
        "material": material,
        "purchased": purchased,
        "production": production + setup,
        "overhead_profit": material_profit + overhead + process_profit,
        "packaging": packaging,
        "freight": freight,
        "tooling": tooling,
    }


def find_label_in_rows(ws, start_row: int, end_row: int, *parts: str, exact=False):
    wanted = {norm(part) for part in parts}
    matches = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            text = norm(cell.value)
            matched = text in wanted if exact else any(part in text for part in wanted)
            if matched:
                matches.append(cell)
    if not matches:
        raise LookupError(f"第 {start_row}:{end_row} 行找不到标签: {parts}")
    return max(matches, key=lambda cell: (cell.column, -cell.row))


def last_numeric_in_column(data_ws, column: int, start_row: int, end_row: int):
    for row in range(end_row, start_row - 1, -1):
        value = data_ws.cell(row, column).value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return value
    raise LookupError(f"第 {column} 列的 {start_row}:{end_row} 行没有数值")


def extract_india1_values(ws, ds) -> dict[str, Any]:
    date_label = find_rightmost_label(ws, "quotation date")
    date_value = next_value_right(ds, ds.cell(date_label.row, date_label.column), 20)

    material_price_header = find_rightmost_label(ws, "raw material price", exact=True)
    purchased_header = find_rightmost_label(
        ws, "purchased parts and external job work", "purchased parts and external services")
    production_section = find_rightmost_label(ws, "production costs", exact=True)
    material_result_col = find_rightmost_label(ws, "cost of material", exact=True).column

    raw_prices = [
        number(ds.cell(row, material_price_header.column).value)
        for row in range(material_price_header.row + 1, purchased_header.row)
    ]
    raw_prices = [value for value in raw_prices if value]
    material_basis = max(raw_prices) if raw_prices else 0
    material = last_numeric_in_column(
        ds, material_result_col, material_price_header.row + 1, purchased_header.row - 1)

    purchased_total_label = find_label_in_rows(
        ws, purchased_header.row, production_section.row - 1, "Σ 1", "Σ 2", exact=True)
    purchased = next_numeric_right(
        ds, ds.cell(purchased_total_label.row, purchased_total_label.column), ws.max_column)

    production_label = find_rightmost_label(ws, "Manufacturing Cost", exact=True)
    production = next_numeric_right(
        ds, ds.cell(production_label.row, production_label.column), ws.max_column)
    production_result_col = find_rightmost_label(ws, "production costs", exact=True).column
    india_overhead_labels = (
        "+ ICC ON RM",
        "+ REJECTION",
        "Profit on RM",
        "OH on tool maintenance",
        "Overhead on Conversion.",
    )
    overhead_profit = sum(
        value_at_label_and_column(ws, ds, (label,), production_result_col)
        for label in india_overhead_labels
    )

    packaging_header = find_rightmost_label(ws, "packaging costs", exact=True)
    packaging_total = find_rightmost_label(ws, "costs2")
    packaging = number(ds.cell(packaging_total.row, packaging_header.column).value)

    tooling_header = find_rightmost_label(ws, "Σ tooling costs", exact=True)
    tooling = last_numeric_in_column(
        ds, tooling_header.column, tooling_header.row + 1, ws.max_row)

    return {
        "date": date_value,
        "material_basis": material_basis,
        "material": material,
        "purchased": purchased,
        "production": production,
        "overhead_profit": overhead_profit,
        "packaging": packaging,
        # 印度汇总模板的 Inland Freight 使用目标表原有公式，不从报价单覆盖。
        "tooling": tooling,
    }


def extract_india_double_region_values(
        ws, ds, overhead_labels: Iterable[str], freight_total_label: str) -> dict[str, Any]:
    """Extract the marked EUR-region values used by the India/India2 templates."""
    date_label = find_rightmost_label(ws, "quotation date")
    date_value = next_value_right(ds, ds.cell(date_label.row, date_label.column), 20)

    material_price_header = find_rightmost_label(ws, "raw material price", exact=True)
    purchased_header = find_rightmost_label(
        ws, "purchased parts and external job work", "purchased parts and external services")
    raw_prices = [
        number(ds.cell(row, material_price_header.column).value)
        for row in range(material_price_header.row + 1, purchased_header.row)
    ]
    raw_prices = [value for value in raw_prices if value]
    material_basis = max(raw_prices) if raw_prices else 0

    material_label = find_rightmost_label(ws, "Σ 1", exact=True)
    purchased_label = find_rightmost_label(ws, "Σ 2", exact=True)
    material = next_numeric_right(
        ds, ds.cell(material_label.row, material_label.column), ws.max_column)
    purchased = next_numeric_right(
        ds, ds.cell(purchased_label.row, purchased_label.column), ws.max_column)

    production_result_col = find_rightmost_label(ws, "production costs", exact=True).column
    production = value_at_label_and_column(
        ws, ds, ("manufaction costs", "manufacturing costs"), production_result_col)
    setup = value_at_label_and_column(ws, ds, ("set up costs",), production_result_col)
    overhead_profit = sum(
        value_at_label_and_column(ws, ds, (label,), production_result_col)
        for label in overhead_labels
    )

    packaging_header = find_rightmost_label(ws, "packaging costs", exact=True)
    packaging_total = find_rightmost_label(ws, "costs2")
    packaging = number(ds.cell(packaging_total.row, packaging_header.column).value)

    freight_header = find_rightmost_label(ws, "shipping costs", exact=True)
    freight_total = find_rightmost_label(ws, freight_total_label)
    freight = number(ds.cell(freight_total.row, freight_header.column).value)

    tooling_header = find_rightmost_label(ws, "Σ tooling costs", exact=True)
    tooling = last_numeric_in_column(
        ds, tooling_header.column, tooling_header.row + 1, ws.max_row)

    return {
        "date": date_value,
        "material_basis": material_basis,
        "material": material,
        "purchased": purchased,
        "production": production + setup,
        "overhead_profit": overhead_profit,
        "packaging": packaging,
        "freight": freight,
        "tooling": tooling,
    }


def extract_india_values(ws, ds) -> dict[str, Any]:
    return extract_india_double_region_values(
        ws,
        ds,
        ("+ overhead", "+ profit on manufcturing process"),
        "Ex-works Cost",
    )


def extract_india2_values(ws, ds) -> dict[str, Any]:
    return extract_india_double_region_values(
        ws,
        ds,
        (
            "+ Power & Fuel",
            "+ ICC",
            "+ overhead",
            "+ profit on manufcturing process",
        ),
        "FOB costs3",
    )


def normalize_country(value: Any) -> str | None:
    text = norm(value)
    if text in {"china", "中国"}:
        return "China"
    if text in {"india", "印度"}:
        return "India"
    if text == "india1":
        return "India1"
    if text == "india2":
        return "India2"
    return None


def extract_values(path: Path, country: str) -> dict[str, Any]:
    formulas = load_workbook(path, data_only=False, read_only=False)
    values = load_workbook(path, data_only=True, read_only=False)
    ws, ds = select_quote_sheet(formulas, values, country)
    if country == "China":
        return extract_china_values(ws, ds)
    if country == "India1":
        return extract_india1_values(ws, ds)
    if country == "India":
        return extract_india_values(ws, ds)
    if country == "India2":
        return extract_india2_values(ws, ds)
    raise ValueError(f"不支持的供应商国家: {country}")


def build_file_index(root: Path, summary: Path) -> dict[str, list[Path]]:
    index: dict[str, list[Path]] = {}
    for path in root.rglob("*"):
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"} and path.resolve() != summary.resolve():
            index.setdefault(norm(path.stem), []).append(path)
    return index


def resolve(index: dict[str, list[Path]], requested: str, search_root: Path) -> Path:
    matches = index.get(norm(Path(requested).stem), [])
    if not matches:
        raise FileNotFoundError(f"递归目录中找不到同名文件: {requested}")
    valid = [path for path in matches if zipfile.is_zipfile(path)]
    if not valid:
        raise RuntimeError("找到同名文件，但都不是有效的 Excel 文件: " + ", ".join(map(str, matches)))
    return min(valid, key=lambda p: (len(p.relative_to(search_root).parts), str(p).lower()))


# PyInstaller 单文件 EXE 会把源码释放到临时目录；onefile 下 sys.executable
# 指向临时解压目录，必须用 sys.argv[0] 定位用户双击的 EXE 所在目录。
script_dir = (Path(sys.argv[0]).resolve().parent if getattr(sys, 'frozen', False)
              else Path(__file__).resolve().parent)


def main() -> int:
    parser = argparse.ArgumentParser(description="按名称从供应商报价表提取数据并回填汇总 xlsm")
    parser.add_argument(
        "summary", type=Path, nargs="?",
        help="汇总工作簿；不传时读取脚本同目录下的 Summary_RFQ comparison.xlsm")
    parser.add_argument("--search-root", type=Path, help="递归搜索目录；默认主工作簿所在目录")
    parser.add_argument("--output", type=Path, help="输出 xlsm；默认输出到脚本/EXE 同目录")
    parser.add_argument("--sheet", action="append", help="只处理指定 sheet，可重复；默认处理全部 sheet")
    parser.add_argument("--log", type=Path, help="JSON 日志路径")
    args = parser.parse_args()

    summary = (args.summary or (script_dir / "Summary_RFQ comparison.xlsm")).resolve()
    if not summary.is_file():
        parser.error(f"找不到汇总工作簿: {summary}\n请把脚本与 Summary_RFQ comparison.xlsm 放在同一目录。")
    search_root = (args.search_root or summary.parent).resolve()
    output = (
        args.output
        or (script_dir / f"{summary.stem}_filled.xlsm")
    ).resolve()
    if summary == output:
        raise ValueError("输出路径不能覆盖源工作簿")
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(summary, keep_vba=True, data_only=False)
    requested_sheets = args.sheet or wb.sheetnames
    index = build_file_index(search_root, summary)
    log = {"summary": str(summary), "output": str(output), "search_root": str(search_root), "items": []}
    updates: dict[str, dict[str, Any]] = {}

    for sheet_name in requested_sheets:
        if sheet_name not in wb.sheetnames:
            log["items"].append({"sheet": sheet_name, "status": "error", "error": "sheet 不存在"})
            continue
        if norm(sheet_name) == "summary":
            continue  # 封面/汇总索引页，不是供应商回填明细页
        ws = wb[sheet_name]
        for col in SUPPLIER_COLUMNS:
            requested = ws.cell(11, col).value
            if not isinstance(requested, str) or not requested.strip():
                continue
            country_value = ws.cell(12, col).value
            country = normalize_country(country_value)
            item = {
                "sheet": sheet_name,
                "column": ws.cell(11, col).column_letter,
                "requested": requested,
                "country": country_value,
            }
            if country is None:
                item.update(
                    status="skipped",
                    error=f"第 12 行模板不是 China、India、India1 或 India2: {country_value!r}",
                )
                log["items"].append(item)
                continue
            try:
                source = resolve(index, requested, search_root)
                extracted = extract_values(source, country)
                for key, value in extracted.items():
                    row = DEFAULT_ROWS[key]
                    coordinate = ws.cell(row, col).coordinate
                    updates.setdefault(sheet_name, {})[coordinate] = value
                item.update(
                    status="ok",
                    normalized_country=country,
                    source=str(source),
                    values={
                        k: str(v) if isinstance(v, datetime) else v
                        for k, v in extracted.items()
                    },
                )
            except FileNotFoundError as exc:
                item.update(status="skipped", error=str(exc))
            except Exception as exc:
                item.update(status="error", error=str(exc))
            log["items"].append(item)

    write_preserving_excel_features(summary, output, updates)
    log_path = (args.log or output.with_suffix(".json")).resolve()
    log_path.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")
    ok = sum(i["status"] == "ok" for i in log["items"])
    skipped = sum(i["status"] == "skipped" for i in log["items"])
    errors = sum(i["status"] == "error" for i in log["items"])
    print(f"完成: {ok} 个文件，跳过: {skipped} 个，失败: {errors} 个；输出: {output}；日志: {log_path}")
    return 1 if errors else 0


if __name__ == "__main__":
    try:
        rc = main()
    except SystemExit:
        raise
    except Exception:
        import traceback
        error_log = script_dir / "fill_sliding_roof_brackets_错误日志.txt"
        try:
            error_log.write_text(traceback.format_exc(), encoding="utf-8")
        except OSError:
            pass
        print(f"❌ 程序出错，详情已写入: {error_log}")
        traceback.print_exc()
        input("程序异常退出，按回车键关闭窗口...")
        sys.exit(1)
    else:
        print("按回车键退出...")
        input()
        sys.exit(rc)
